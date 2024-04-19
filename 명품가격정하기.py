import tkinter
import pandas as pd
import re
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import os
import tkinter as tk
from tkinter import filedialog, simpledialog, Radiobutton, IntVar, Listbox , Entry, Button, Listbox, Scrollbar,StringVar



def rename_columns_by_exact_match(df, keywords=['model','category', 'brand', 'cost', 'retail']):
    """
    DataFrame의 각 열을 확인하여 주어진 단어들 중 하나가 정확히 존재하면 해당 열의 이름을 그 단어로 변경합니다.
    만약 'cost'나 'retail' 중 하나를 찾으면, 다른 하나는 더 이상 찾지 않습니다.
    :param df: DataFrame
    :param keywords: 찾을 단어 목록
    :return: 열 이름이 변경된 DataFrame
    """
    found_cost_or_retail = False
    for col in df.columns:
        for keyword in keywords:
            # 이미 'cost' 또는 'retail' 중 하나를 찾았다면, 다른 하나는 더 이상 찾지 않음
            if found_cost_or_retail and keyword in ['cost', 'retail']:
                continue

            # 열의 셀 값 중에서 해당 키워드와 정확히 일치하는 경우를 찾습니다.
            if df[col].astype(str).str.fullmatch(keyword, case=False, na=False).any():
                df.rename(columns={col: keyword}, inplace=True)
                if keyword in ['cost', 'retail']:
                    found_cost_or_retail = True  # 'cost' 또는 'retail'을 찾음
                break  # 해당 단어를 찾으면 다음 열로 넘어갑니다
    return df






def find_cost_or_retail_column(df):
    cost_pattern = re.compile(r'\bcost\b', re.IGNORECASE)
    retail_pattern = re.compile(r'\bretail\b', re.IGNORECASE)

    for col in df.columns:
        col_str = str(col)  # 열 이름을 문자열로 변환
        if "Unnamed:" in col_str:  
            for index, value in df[col].items():
                value_str = str(value)
                if cost_pattern.search(value_str):  
                    df.rename(columns={col: 'Cost'}, inplace=True)
                    print(f"Column '{col}' renamed to 'Cost'")
                    return 'Cost', 'cost'
                elif retail_pattern.search(value_str):  
                    df.rename(columns={col: 'Retail'}, inplace=True)
                    print(f"Column '{col}' renamed to 'Retail'")
                    return 'Retail', 'retail'

    print("Neither 'cost' nor 'retail' column found in the DataFrame.")
    return None, None






def clean_value(value):
    # 값이 비어있는지 확인
    if pd.isna(value) or value == "" or value == "nan":
        print(f"Empty or invalid value found: {value}")
        return None

    # 값이 비어있지 않은 경우, 숫자로 변환
    try:
        cleaned_value = re.sub(r'[^\d.,]', '', str(value))
        cleaned_value = float(cleaned_value.replace(',', '.'))
        print(f"Original value: {value}, Cleaned value: {cleaned_value}")
        return cleaned_value
    except ValueError as e:
        print(f"Error converting value: {value}, Error: {e}")
        return None




# 브랜드 열과 시작 행 찾기
def find_brand_column_and_row(df):
    for col in df.columns:
        for i, cell in enumerate(df[col].astype(str)):
            if 'brand' in cell.lower():
                return col, i + 1
    return None, None

# 'model' 열 찾기 (대소문자 구분 없이)
def find_model_column(df):
    for col in df.columns:
        if df[col].astype(str).str.contains('model', case=False, na=False).any():
            return col
    return None

def extract_unique_brands(df, brand_col, start_row):
    unique_brands = df[brand_col].iloc[start_row+1:].dropna().unique()
    #print(f"Extracted brands: {unique_brands}")  # 추출된 브랜드 출력
    return unique_brands

def update_selected_brands(brand, var, selected_brands):
    if var.get() == 1:
        print(f"{brand} selected")
        if brand not in selected_brands:
            selected_brands.append(brand)
    else:
        print(f"{brand} deselected")
        if brand in selected_brands:
            selected_brands.remove(brand)

def find_category_column(df):
    for col in df.columns:
        if df[col].astype(str).str.contains('category', case=False, na=False).any():
            return col
    return None

def extract_unique_categories(df, brand_col, category_col, selected_brands):
    unique_categories_by_brand = {}
    for brand in selected_brands:
        unique_categories = df[df[brand_col] == brand][category_col].unique()
        unique_categories_by_brand[brand] = unique_categories
    return unique_categories_by_brand


def ask_for_brand_conditions(brands, categories_by_brand):
    root = tk.Tk()
    root.title("Enter Conditions for Each Brand and Category")

    entries = {}  # 브랜드와 카테고리별 입력 필드를 저장할 딕셔너리

    # 브랜드와 카테고리별로 레이블과 입력 필드 생성
    for brand in brands:
        tk.Label(root, text=brand).pack()
        for category in categories_by_brand[brand]:
            frame = tk.Frame(root)
            frame.pack()
            tk.Label(frame, text=category).pack(side=tk.LEFT)
            entry = tk.Entry(frame)
            entry.pack(side=tk.RIGHT)
            entries[(brand, category)] = entry

    brand_conditions = {}  # 입력된 조건을 저장할 딕셔너리

    # 사용자 입력 데이터를 처리하는 함수
    def on_ok():
        for brand in brands:
            conditions = {}
            for category in categories_by_brand[brand]:
                entry = entries[(brand, category)]
                condition_value = float(entry.get())
                conditions[category] = condition_value
            brand_conditions[brand] = conditions
        root.quit()

    # 확인 버튼
    ok_button = tk.Button(root, text="OK", command=on_ok)
    ok_button.pack()

    root.mainloop()
    root.destroy()  # 창 닫기

    return brand_conditions


def ask_for_conditions_exchange_rate_margin_iva(brands):
    def update_listbox(listbox, items):
        listbox.delete(0, tk.END)
        for item in items:
            listbox.insert(tk.END, item)

    def add_selected_brand():
        selected_brand = brand_listbox.get(tk.ANCHOR)
        if selected_brand and selected_brand not in selected_brands:
            selected_brands.append(selected_brand)
            print(f"{selected_brand} added to selected brands")
            update_listbox(selected_brands_listbox, selected_brands)

    selected_brands = []
    exchange_rate = 0
    margin = 0

    root = tk.Tk()
    root.title("Select Brands for Calculation")

    iva_var = IntVar(root, value=2)

    brand_listbox = Listbox(root)
    brand_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar = Scrollbar(root, orient="vertical", command=brand_listbox.yview)
    scrollbar.pack(side=tk.LEFT, fill=tk.Y)
    brand_listbox.config(yscrollcommand=scrollbar.set)

    # 선택된 브랜드를 위한 리스트박스와 스크롤바 설정
    selected_brands_frame = tk.Frame(root)
    selected_brands_frame.pack()

    selected_brands_listbox = Listbox(selected_brands_frame, height=6)  # height 값을 조정하여 크기 변경
    selected_brands_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    selected_brands_scrollbar = Scrollbar(selected_brands_frame, orient="vertical")
    selected_brands_scrollbar.pack(side=tk.LEFT, fill=tk.Y)

    # 리스트박스와 스크롤바를 서로 연결
    selected_brands_listbox.config(yscrollcommand=selected_brands_scrollbar.set)
    selected_brands_scrollbar.config(command=selected_brands_listbox.yview)


    def add_all_brands():
        for brand in brands:
            if brand not in selected_brands:
                selected_brands.append(brand)
        update_listbox(selected_brands_listbox, selected_brands)
        print("All brands added")

    def delete_selected_brand():
        selected_brand = selected_brands_listbox.get(tk.ANCHOR)
        if selected_brand in selected_brands:
            selected_brands.remove(selected_brand)
            update_listbox(selected_brands_listbox, selected_brands)
            print(f"{selected_brand} removed from selected brands")

    # 버튼 프레임을 먼저 배치합니다.
    button_frame = tk.Frame(root)
    button_frame.pack() 

    # 'All' 버튼 추가
    all_button = Button(button_frame, text="All", command=add_all_brands)
    all_button.pack(side=tk.LEFT)

    # 'Add' 버튼 추가
    add_button = Button(button_frame, text="Add", command=add_selected_brand)
    add_button.pack(side=tk.LEFT)

    # 'Delete' 버튼 추가
    delete_button = Button(button_frame, text="Delete", command=delete_selected_brand)
    delete_button.pack(side=tk.LEFT)
    
    update_listbox(brand_listbox, brands)

    tk.Label(root, text="Exchange Rate:").pack()
    exchange_rate_entry = tk.Entry(root)
    exchange_rate_entry.pack()

    tk.Label(root, text="Margin (%):").pack()
    margin_entry = tk.Entry(root)
    margin_entry.pack()

    iva_included_button = Radiobutton(root, text="IVA Included", variable=iva_var, value=1)
    iva_excluded_button = Radiobutton(root, text="IVA Excluded", variable=iva_var, value=0)
    no_retail_button = Radiobutton(root, text="No Retail", variable=iva_var, value=2)
    no_retail_button.select()

    iva_included_button.pack()
    iva_excluded_button.pack()
    no_retail_button.pack()

    def get_values():
        nonlocal exchange_rate, margin
        exchange_rate = float(exchange_rate_entry.get())
        margin = float(margin_entry.get()) / 100 + 1
        root.quit()

    ok_button = tk.Button(root, text="OK", command=get_values)
    ok_button.pack()

    cancel_button = tk.Button(root, text="Cancel", command=root.quit)
    cancel_button.pack()

    root.mainloop()
    root.destroy()  # 창 닫기

    iva_option = iva_var.get()
    print("IVA option selected:", iva_option)

    return selected_brands, exchange_rate, margin, iva_option

#계산기
def calculate_values(value, exchange_rate, brand_conditions, brand,category ,margin, iva_option, cost_or_retail, trade_method, column_name=None):
    print(f"Current trade method in calculate_values: {trade_method}")  # 디버깅 코드 추가
    condition = brand_conditions.get(brand, {}).get(category, 0) / 100
    result = None
    calculation_method = ""


    if trade_method == "EXW":
        print("Performing EXW calculation")
        if cost_or_retail == 'cost':
            result = value * (1+condition) * exchange_rate * margin
            calculation_method = "EXW Cost Calculation"
        elif cost_or_retail == 'retail':
            if iva_option == 1:
                result = (value / 1.22) * (1 - condition) * exchange_rate * margin
                calculation_method = "EXW Retail with IVA Calculation"
            elif iva_option == 0:
                result = value * (1 - condition) * exchange_rate * margin
                calculation_method = "EXW Retail without IVA Calculation"

    elif trade_method == "DDP":
        # DDP 계산 로직
        print("Performing DDP calculation")
        if cost_or_retail == 'cost':
            if column_name == "FTA 가능 DDP":
                multiplier = 1.1
            elif column_name == "FTA 불가 DDP":
                multiplier = 1.3
            elif column_name == "FTA 불가 DDP (신발,의류)":
                multiplier = 1.4
            else:
                multiplier = 1.0

            multiplier *= (condition + 1)
            result = value * margin * exchange_rate * multiplier
            calculation_method = "DDP Cost Calculation"
        
        elif cost_or_retail == 'retail':
            if iva_option == 1:
                if column_name == "FTA 가능 DDP":
                    multiplier = 1.1 * margin
                elif column_name == "FTA 불가 DDP":
                    multiplier = 1.3 * margin
                elif column_name == "FTA 불가 DDP (신발,의류)":
                    multiplier = 1.4 * margin
                else:
                    multiplier = 1.0

                result = (value / 1.22) * (1 - condition) * exchange_rate * multiplier
                calculation_method = "DDP Retail with IVA Calculation"
            elif iva_option == 0:
                if column_name == "FTA 가능 DDP":
                    multiplier = 1.1 * margin
                elif column_name == "FTA 불가 DDP":
                    multiplier = 1.38 * margin
                elif column_name == "FTA 불가 DDP (신발,의류)":
                    multiplier = 1.4 * margin
                else:
                    multiplier = 1.0

                result = value * (1 - condition) * exchange_rate * multiplier
                calculation_method = "DDP Retail without IVA Calculation"
        if result is not None:
            # 결과값을 소수점 세 자리까지 반올림
            result = round(result, 3)

    
    # 결과와 계산 방법을 튜플로 반환
    print(f"Calculated result: {result}, Calculation Method: {calculation_method}")
    return str(result), calculation_method


def select_trade_method():
    root = tk.Tk()
    root.title("Select Trade Method")

    # 라디오 버튼을 위한 변수
    trade_method_var = tk.StringVar(value="DDP")

    # DDP와 EXW를 위한 라디오 버튼
    ddp_radio = tk.Radiobutton(root, text="DDP", variable=trade_method_var, value="DDP")
    exw_radio = tk.Radiobutton(root, text="EXW", variable=trade_method_var, value="EXW")

    ddp_radio.pack()
    exw_radio.pack()

    # 사용자 선택을 저장할 변수
    selected_method = [None]

    def on_ok():
        selected_method[0] = trade_method_var.get()
        print(f"Selected trade method: {selected_method[0]}")
        root.destroy()

    ok_button = tk.Button(root, text="OK", command=on_ok)
    ok_button.pack()

    root.mainloop()

    # 사용자의 선택을 반환
    return selected_method[0]

def add_column_titles(ws, title_row,start_col, titles,fill):
    """
    지정된 시작 열부터 시작하여 제목을 추가합니다.
    :param ws: 워크시트 객체
    :param start_col: 제목을 추가할 시작 열의 인덱스
    :param titles: 추가할 제목의 리스트
    """
    for i, title in enumerate(titles, start=start_col):
        cell = ws.cell(row=title_row, column=i, value=title)
        cell.fill = fill


def process_excel(file_path, df, brand_conditions, exchange_rate, margin, brand_col, start_row, iva_option, selected_brands, trade_method):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 'cost'나 'retail' 열의 위치를 찾습니다.
    cost_or_retail_col_name = None
    if 'cost' in df.columns:
        cost_or_retail_col_name = 'cost'
    elif 'retail' in df.columns:
        cost_or_retail_col_name = 'retail'

    if cost_or_retail_col_name is None:
        print("Neither 'cost' nor 'retail' column found in the DataFrame.")
        return

    cost_or_retail_col_index = df.columns.get_loc(cost_or_retail_col_name) + 1  # 열 인덱스는 1부터 시작


    # 색상 정의
    dark_green_fill = PatternFill(start_color='006400', end_color='006400', fill_type='solid')
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    title_row = None
    for row in ws.iter_rows(min_row=1, max_col=cost_or_retail_col_index, max_row=ws.max_row):
        for cell in row:
            if cell.value in ['cost', 'retail']:
                title_row = cell.row
                break
        if title_row:
            break

    if title_row is None:
        print("No 'cost' or 'retail' row found.")
        return


    # 새 열 추가
    ws.insert_cols(cost_or_retail_col_index + 1, amount=3)
    titles = ['FTA 가능 DDP', 'FTA 불가 DDP', 'FTA 불가 DDP (신발,의류)']
    add_column_titles(ws, title_row, cost_or_retail_col_index + 1, titles, dark_green_fill)

    # 원본 데이터 이동
    for col in range(ws.max_column, cost_or_retail_col_index + 3, -1):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=col).value = ws.cell(row=row, column=col - 3).value

    # 새로운 데이터 삽입
    for i, row in enumerate(df.iterrows(), start=0):
        row_data = row[1]
        brand = row_data[brand_col]
        if brand in selected_brands:
            category_col = find_category_column(df)
            if category_col is None:
                continue

            category = row_data[category_col]
            raw_value = row_data[cost_or_retail_col_name]
            cleaned_value = clean_value(raw_value)

            if cleaned_value is not None:
                processed_values = []
                for col_name in ['FTA 가능 DDP', 'FTA 불가 DDP', 'FTA 불가 DDP (신발,의류)']:
                    processed_value, _ = calculate_values(cleaned_value, exchange_rate, brand_conditions, brand, category, margin, iva_option, cost_or_retail_col_name, trade_method, col_name)
                    processed_values.append(processed_value)

                # 계산된 값을 해당 행의 새 열에 삽입
            for j, value in enumerate(processed_values, start=cost_or_retail_col_index + 1):
                cell = ws.cell(row=i + 1, column=j, value=value)
                cell.fill = light_green_fill



    wb.save(file_path)
    print(f"Data updated and saved in {file_path}")

        
def main():
    # 1. 무역 방식 선택
    trade_method = select_trade_method()
    print(f"Selected trade method: {trade_method}")

    # 2. 엑셀 파일 선택
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title="Select an Excel file", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path:
        return

    df = pd.read_excel(file_path, header=None)
    df = rename_columns_by_exact_match(df)
    print(df.columns)

    # 3. 브랜드 선택 및 환율, 마진 입력
    brand_col, start_row = find_brand_column_and_row(df)
    if brand_col is None:
        print("No 'brand' column found")
        return

    brands = extract_unique_brands(df, brand_col, start_row)
    if len(brands) == 0:
        print("No brands selected.")
        return

    selected_brands, exchange_rate, margin, iva_option = ask_for_conditions_exchange_rate_margin_iva(brands)
    print(f"IVA option selected: {iva_option}")

    if len(selected_brands) == 0:
        print("No brands selected for calculation.")
        return

    # 4. 브랜드별 카테고리 조건 입력
    category_col = find_category_column(df)
    unique_categories_by_brand = extract_unique_categories(df, brand_col, category_col, selected_brands)
    brand_conditions = ask_for_brand_conditions(selected_brands, unique_categories_by_brand)

    # 5. 새 엑셀 파일 이름 입력 및 저장
    processed_df = process_excel(file_path,df, brand_conditions, exchange_rate, margin, brand_col, start_row, iva_option, selected_brands, trade_method)

    if processed_df is not None:
        print("Data processing completed.")

if __name__ == "__main__":
    main()
